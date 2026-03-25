import os
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# --- 설정 (환경에 맞게 수정) ---
BASE_PATH = r"Z:\"
EXCEL_FILE = r"Z:\WorkSheet.xlsx"
OUTPUT_FILE = r"Z:\WorkSheet_Result.xlsx"

# 셀의 Width/Height 보다 작게 삽입하기 위한 최대 픽셀 크기 설정
# 실제 Excel의 FIB1, FIB2 셀 크기에 맞춰 값을 조정하십시오.
MAX_IMG_WIDTH = 120  
MAX_IMG_HEIGHT = 80  

def process_images_to_excel():
    # 1. 엑셀 파일 로드 및 헤더 컬럼 식별
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # 첫 번째 행을 기준으로 컬럼 이름과 인덱스 매핑 (1-based)
    header = {cell.value: cell.column for cell in ws[1]}
    
    col_dqm = header.get("DQM")
    col_mount = header.get("MOUNT")
    col_fib1 = header.get("FIB1")
    col_fib2 = header.get("FIB2")

    if not all([col_dqm, col_mount, col_fib1, col_fib2]):
        print("Error: 엑셀 파일에서 필수 컬럼(DQM, MOUNT, FIB1, FIB2)을 찾을 수 없습니다.")
        return

    # 2. 모든 행에 대해 Loop (2행부터 마지막 행까지)
    for row in range(2, ws.max_row + 1):
        dqm_val = ws.cell(row=row, column=col_dqm).value
        mount_val = ws.cell(row=row, column=col_mount).value

        if not dqm_val or not mount_val:
            continue

        # 3. FOLDER_TREE 식별
        folder_path = os.path.join(BASE_PATH, str(dqm_val))
        if not os.path.isdir(folder_path):
            continue

        # 4. 해당 폴더 내의 {MOUNT}_*.tif 파일 검색 및 SEQ 파싱
        target_prefix = f"{mount_val}_"
        matched_files = []
        
        for file_name in os.listdir(folder_path):
            if file_name.startswith(target_prefix) and file_name.lower().endswith('.tif'):
                # '{MOUNT}_' 이후부터 '.tif' 이전까지의 문자열을 추출하여 정수로 변환
                seq_str = file_name[len(target_prefix):-4]
                try:
                    seq_num = int(seq_str)
                    matched_files.append((seq_num, file_name))
                except ValueError:
                    pass # SEQ가 숫자가 아닌 경우 무시

        if not matched_files:
            continue

        # 숫자형 SEQ를 기준으로 오름차순 정렬
        matched_files.sort(key=lambda x: x[0])

        # 5 & 6. 조건에 맞는 파일 선택
        # 2번째로 작은 숫자 (인덱스 1), 가장 큰 숫자 (인덱스 -1)
        fib1_file = matched_files[1][1] if len(matched_files) >= 2 else None
        fib2_file = matched_files[-1][1] if len(matched_files) >= 1 else None

        # 이미지 삽입 내부 함수
        def insert_image(filename, col_idx):
            if not filename: 
                return
            
            img_path = os.path.join(folder_path, filename)
            try:
                with PILImage.open(img_path) as img:
                    # Cell Width/Height 내에 들어가도록 비율 유지 축소
                    img.thumbnail((MAX_IMG_WIDTH, MAX_IMG_HEIGHT))
                    
                    # 엑셀 호환성을 위해 메모리 상에서 PNG 형식으로 변환
                    img_byte_arr = BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    
                    xl_img = OpenpyxlImage(img_byte_arr)
                    cell_address = ws.cell(row=row, column=col_idx).coordinate
                    ws.add_image(xl_img, cell_address)
            except Exception as e:
                print(f"이미지 처리 중 오류 발생 ({img_path}): {e}")

        insert_image(fib1_file, col_fib1)
        insert_image(fib2_file, col_fib2)

    # 결과 저장 (원본 보호를 위해 새 파일로 저장)
    wb.save(OUTPUT_FILE)
    print(f"작업 완료. 결과가 {OUTPUT_FILE}에 저장되었습니다.")

if __name__ == "__main__":
    process_images_to_excel()
